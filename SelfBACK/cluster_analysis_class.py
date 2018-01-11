# -*- coding: utf-8 -*-
import numpy as np
import copy as cp
import matplotlib.cm as cm
import matplotlib.pyplot as plt 

from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_samples, silhouette_score
from openpyxl import Workbook
from collections import Counter
import utilities as tools
from kmodes.kmodes import KModes
from kmodes.kprototypes import KPrototypes




class dataset_clustering():
    
    def __init__(self, data_set=[], data_for_clustering=[], attributes_for_clustering=[], categoricals=[], ordinals=[], n_test_clusters=1, clustering_method='KMeans'):
    
        self.n_test_clusters = n_test_clusters
        self.data_set = data_set 
        self.data_for_clustering = data_for_clustering
        self.attributes_for_clustering = attributes_for_clustering
        self.categoricals = categoricals
        self.ordinals = ordinals
        self.clustering_method = clustering_method

##_______________________  K means Clustering ____________________________________

    def run_KMeans (self):
        
        K_number = []
        squared_distance = []
        
        self.Kmean_input_data = self.data_for_clustering[1:, 1:]
        
        for n_clusters in range(2,self.n_test_clusters+1):
                    
            fig, ax1 = plt.subplots(1)
            fig.set_size_inches(6,9)
            ax1.set_xlim([-0.2,0.6])
            
        # The (n_clusters+1)*10 is for inserting blank space between silhouette
        # plots of individual clusters, to demarcate them clearly.
            ax1.set_ylim([0, len(self.Kmean_input_data) + (n_clusters+1)*10])
            
            clusterer = KMeans(n_clusters = n_clusters, init='k-means++', n_init=25, max_iter = 1000)
            cluster_labels = clusterer.fit_predict(self.Kmean_input_data)
            K_number.append(n_clusters)
            squared_distance.append(clusterer.inertia_)
            
            
            silhouette_avg = silhouette_score(self.Kmean_input_data, cluster_labels)
            print('\nHaving', n_clusters, 'clusters, the average silhouette score is: ', silhouette_avg)
           
            
            sample_silhouette_values = silhouette_samples(self.Kmean_input_data, cluster_labels)
            
            y_lower = 10
            
            for j in range(n_clusters):
                ith_cluster_silhouette_values = sample_silhouette_values[cluster_labels == j]
                
                ith_cluster_silhouette_values.sort()
                
                size_cluster_j = ith_cluster_silhouette_values.shape[0]
                
                y_upper = y_lower + size_cluster_j
                color = cm.spectral(float(j) / n_clusters)
                ax1.fill_betweenx(np.arange(y_lower, y_upper), 0, ith_cluster_silhouette_values, facecolor=color, edgecolor=color, alpha=0.7)
                
                ax1.text(-0.05, y_lower + 0.5 * size_cluster_j, str(j+1))
                
        # Compute the new y_lower for next plot    
                y_lower= y_upper +10
            
            ax1.set_title('The silhouette plot for %s clusters' %n_clusters)
            ax1.set_xlabel('The silhouette coeficient values')
            ax1.set_ylabel('cluster label')
            
            ax1.axvline(x=silhouette_avg, color='red', linestyle='--')
            
            ax1.set_yticks([])
            ax1.set_xticks([-0.2, 0, 0.2, 0.4, 0.6, 0.8, 1])
            
            fig.savefig('./results/plots/SilhouettePlot-' + str(n_clusters) + 'clusters.png', bbox_inches='tight', dpi=150) #
            #plt.show()
            plt.close()
            
    
            
        max_squared_distance = np.amax(squared_distance)
        min_squared_distance = np.amin(squared_distance)
        fig, (ax1) = plt.subplots(1)
        fig.set_size_inches(6,9)
        plt.plot(K_number, squared_distance, 'ro')
        plt.xlabel('Number of Clusters')
        plt.ylabel('sum of squared distances to the closest centroid')
        plt.axis([0, 11, min_squared_distance-10, max_squared_distance+10])
        fig.savefig('./results/plots/squared distances plot.png', bbox_inches='tight', dpi=150) #
        plt.close()
        #plt.show()
            
        num_of_clusters = input('\n \nFirst check the plots in the folder "/results/plots/" and then enter a suitable K (number of clusters) based on the plots:')
        self.num_of_clusters = int(num_of_clusters)
           
        clusterer = KMeans(n_clusters = self.num_of_clusters, init='k-means++', n_init=20, max_iter = 1000).fit(self.Kmean_input_data)
        
        cluster_column = list(map(int, clusterer.labels_+1))
        cluster_column.insert(0, 'clusters')
        
        
        self.clustered_data = np.column_stack((self.data_for_clustering, cluster_column))
        centroids = clusterer.cluster_centers_
        
    #    for i in range (self.num_of_clusters):
    #        print('\n centroid of cluster', i+1, ':', centroids[i])
            
        
        
        self.cluster_column_index = list(self.clustered_data[0]).index('clusters')
        
        clustered_data_sorted = sorted(self.clustered_data[1:], key=lambda x: x[self.cluster_column_index ])
        clustered_data_sorted = np.vstack((self.clustered_data[0], clustered_data_sorted))
        self.list_of_IDs = clustered_data_sorted[1:, 0]
        column_for_num_of_data = self.adding_num_of_data()
        self.clustered_reduced_dataset_sorted = np.column_stack((clustered_data_sorted, column_for_num_of_data))
        
        return self.clustered_reduced_dataset_sorted
        
        
    
    
    
    def run_KModes(self):
        
        self.Kmodes_input_data = self.data_for_clustering[1:, 1:]
        evaluation_metrics = []
        
        for n_clusters in range(2,self.n_test_clusters):
            self.num_of_clusters = n_clusters
            km = KModes(n_clusters=n_clusters, init='Huang', n_init=50, verbose=False)
            clusters = km.fit_predict(self.Kmodes_input_data)
            cluster_column = list(map(int, km.labels_+1))
            clustered_dataset = np.column_stack((self.Kmodes_input_data, cluster_column))
            self.clustered_reduced_dataset_sorted = np.asarray(sorted(clustered_dataset, key=lambda x: x[-1]))
            
            evaluation_metric = self.scatter_criteria(clustered_dataset = self.clustered_reduced_dataset_sorted, attributes=self.attributes_for_clustering,
                                                         ordinals=[], categoricals=self.attributes_for_clustering)
            evaluation_metrics.append(evaluation_metric)
        
        for n, element in enumerate(evaluation_metrics):
            print('\n>> Having %d clusters:\n  - within cluster scatter (Trace of W): %f \n  - between cluster scatter (Trace of B): %f  \n ' % (n+2, element[0][0], element[0][1]))
            
        num_of_clusters = input('\n \nCheck the results and then enter a suitable K (number of clusters):')
        self.num_of_clusters = int(num_of_clusters)
        
        km = KModes(n_clusters=self.num_of_clusters, init='Huang', n_init=50, verbose=False)
        clusters = km.fit_predict(self.Kmodes_input_data)
        cluster_column = list(map(int, km.labels_+1))
        cluster_column.insert(0, 'clusters')
        clustered_dataset = np.column_stack((self.data_for_clustering, cluster_column))
        self.clustered_reduced_dataset_sorted = np.asarray(sorted(clustered_dataset[1:], key=lambda x: x[-1]))
        self.list_of_IDs = self.clustered_reduced_dataset_sorted[1:, 0]
        column_for_num_of_data = self.adding_num_of_data()
        self.clustered_reduced_dataset_sorted = np.column_stack((self.clustered_reduced_dataset_sorted, column_for_num_of_data))
        
        return self.clustered_reduced_dataset_sorted
    
    
    
    
    
    
    def run_KPrototypes(self):
        
        self.Kprototypes_input_data = self.data_for_clustering[1:, 1:]
        evaluation_metrics = []
        self.cateoricals_indices = []
        categoricals = []
        ordinals = []
        for attribute in self.attributes_for_clustering:
            if attribute in self.categoricals:
                categoricals.append(attribute)
                cateorical_index = list(self.attributes_for_clustering).index(attribute)
                self.cateoricals_indices.append(cateorical_index)
            elif attribute in self.ordinals:
                ordinals.append(attribute)
                
                
        for n_clusters in range(2,self.n_test_clusters+1):
            self.num_of_clusters = n_clusters
            kpro = KPrototypes(n_clusters=n_clusters, init='Cao', n_init=50, verbose=False)
            clusters = kpro.fit_predict(self.Kprototypes_input_data, categorical=self.cateoricals_indices)
            cluster_column = list(map(int, kpro.labels_+1))
            clustered_dataset = np.column_stack((self.Kprototypes_input_data, cluster_column))
            self.clustered_reduced_dataset_sorted = np.asarray(sorted(clustered_dataset, key=lambda x: x[-1]))
            
            evaluation_metric = self.scatter_criteria(clustered_dataset=self.clustered_reduced_dataset_sorted, attributes=self.attributes_for_clustering,
                                                         ordinals=ordinals, categoricals=categoricals)
            evaluation_metrics.append(evaluation_metric)
        
            for n, element in enumerate(evaluation_metrics):
                
                print('\n>> Having %d clusters:\n  - within cluster scatter (Trace of W): %f \n  - between cluster scatter (Trace of B): %f  \n ' % (n+2, element[0][0], element[0][1]))
            
        num_of_clusters = input('\n \nCheck the results and then enter a suitable K (number of clusters):')
        self.num_of_clusters = int(num_of_clusters)
        
        kpro = KPrototypes(n_clusters=self.num_of_clusters, init='Cao', n_init=50, verbose=False)
        clusters = kpro.fit_predict(self.Kprototypes_input_data, categorical=self.cateoricals_indices)
        cluster_column = list(map(int, kpro.labels_+1))
        cluster_column.insert(0, 'clusters')
        clustered_dataset = np.column_stack((self.data_for_clustering, cluster_column))
        clustered_reduced_dataset_sorted = np.asarray(sorted(clustered_dataset[1:], key=lambda x: x[-1]))
        self.clustered_reduced_dataset_sorted = np.vstack((clustered_dataset[0], clustered_reduced_dataset_sorted))
        self.list_of_IDs = self.clustered_reduced_dataset_sorted[1:, 0]
        column_for_num_of_data = self.adding_num_of_data()
        self.clustered_reduced_dataset_sorted = np.column_stack((self.clustered_reduced_dataset_sorted, column_for_num_of_data))
        
        return self.clustered_reduced_dataset_sorted
    
    
    
    
    
    
    
    def scatter_criteria(self, clustered_dataset, attributes, ordinals=[], categoricals=[]):
        
        between_scatter =[]
        within_scatter=[]
        W_matrix_maker =[]
        
        divided_clusters =np.split(clustered_dataset, np.where(np.diff(clustered_dataset[:,-1]))[0]+1)
        
        n_instances = len(clustered_dataset[:,0])
        avg_n_members = n_instances / self.num_of_clusters
        overall_means = tools.mean_values(clustered_dataset, attributes, ordinals, categoricals)
        counter1 = 0
        singular_counter = 0
        
        for cluster in divided_clusters:
            n_cluster_member = len(cluster[:,0])
            cluster_means = tools.mean_values(cluster, attributes, ordinals, categoricals)
            means_diff = list(map(abs, tools.mixed_difference(cluster_means, overall_means, attributes, categoricals))) 
            B_matrix_maker = np.outer(means_diff, means_diff)
            B_matrix_maker = np.multiply(B_matrix_maker, 1/n_cluster_member)
            between_scatter = np.add(between_scatter, B_matrix_maker) if counter1 != 0 else B_matrix_maker
            counter2 = 0
            for instance in cluster:
                instance_diff = list(map(abs, tools.mixed_difference(instance[:-1], cluster_means, attributes, categoricals))) 
                W_matrix_maker_inner = np.outer(instance_diff, instance_diff)
                W_matrix_maker = np.add(W_matrix_maker, W_matrix_maker_inner) if counter2 != 0 else W_matrix_maker_inner
                counter2 += 1
                
            W_matrix_maker = np.multiply((1/n_cluster_member), W_matrix_maker)
            
            within_scatter = np.add(within_scatter, W_matrix_maker) if counter1 != 0 else W_matrix_maker
            
            counter1 += 1
        
        between_scatter = np.multiply(between_scatter, avg_n_members)
        within_scatter = np.multiply(within_scatter, avg_n_members)
        
        trace_W = np.matrix.trace(within_scatter)
        trace_B = np.matrix.trace(between_scatter)
        
#        if np.where(~within_scatter.any(axis=1))[0]: singular_counter =1 
#        zero_row = np.where(~within_scatter.any(axis=1))[0]
#        while zero_row:
#            within_scatter[zero_row[0], zero_row[0]] = 1
#            zero_row = np.where(~within_scatter.any(axis=1))[0]
#        if np.where(~within_scatter.any(axis=1))[0]: print(np.where(~within_scatter.any(axis=1))[0])
#        
#        Winv_B = np.matmul( np.linalg.inv(within_scatter) , between_scatter)
#        
#        trace_Winv_B = np.matrix.trace( Winv_B)
        
        evaluation_metric = [trace_W, trace_B]
        
              
        return evaluation_metric, singular_counter #
    
    
    
    
    
    
    
    
    
    
    
    
    def draw_boxplots(self):
        
        requested_cluster_index = 0
        print ('\n Moreover, you can check the boxplot of all attributes in each cluster.'
               ' To do so ...')
        
        
        while requested_cluster_index !='Q':
            print('\n \nWe have ', self.num_of_clusters, 'clusters.')
            requested_cluster_index = input('Please enter the cluster index which you want to see the boxplot about. If you want to quit, press Q: ')
            
            data_for_boxplot = []
            cluster_column_index = list(self.clustered_reduced_dataset_sorted[0]).index('clusters')
            for row in self.clustered_reduced_dataset_sorted[1:]:
               if  (requested_cluster_index != 'Q' and (int(requested_cluster_index)) == int(row[cluster_column_index ])):
                   data_for_boxplot.append(row)
              
            
            if requested_cluster_index != 'Q':
                
                data_for_boxplot = np.array(data_for_boxplot)
                num_of_instances = len(data_for_boxplot[:,0])
                
                fig = plt.figure(1, figsize=(14,6))
                ax = fig.add_subplot(111)
                
                length = len(data_for_boxplot[0])
                boxplot_data = data_for_boxplot[:,1]
                
                for i in range(2,length-2):
                    boxplot_data = np.vstack((list(boxplot_data), list(data_for_boxplot[:,i])))
                
                boxplot_data = list(boxplot_data)
                ax.boxplot(boxplot_data)
                ax.set_xticklabels(self.attributes_for_clustering)
                ax.set_title('The boxplot for cluster %s' %requested_cluster_index)
                plt.xlabel('Attributes used for clustering', fontsize = 12)
                plt.ylabel('Normalized values of attributes for cluster members', fontsize = 12)
                #plt.show()
                fig.savefig('./results/plots/boxplot-cluster' + requested_cluster_index + '.png', bbox_inches='tight', dpi=150)
                print('Check the plots in the folder "results/plots/" !')
                plt.close()
                #print('Number of instances in this cluster is:', num_of_instances)
                
                
                
                
                
    
    
    def adding_num_of_data(self):
        num_of_IDs_column = ['Num of data']
        for ID in self.list_of_IDs:
            ID_index = list(self.data_set[1:,0]).index(ID)
            num_of_IDs = sum(x is not None for x in self.data_set[ID_index])
            num_of_IDs_column.append(num_of_IDs)
        
        return num_of_IDs_column
    
    
    
    
    
    
    def select_representative_IDs(self,first_results=[], n_iterations=0):
        
        wb_clusters = Workbook()
        wb_representatives = Workbook()
        ws = {i:wb_clusters.create_sheet("cluster",i) for i in range(self.num_of_clusters)}
        ws_representatives = wb_representatives.create_sheet("Representatives",0)
        dest_filename0 = './results/Divided clusters for several runs.xlsx'
        dest_filename1 = './results/Representative patients.xlsx'
    
        
        IDs_of_a_cluster_in_different_runs ={i:[] for i in range(n_iterations)}
        clusters_in_different_runs ={j:dict(IDs_of_a_cluster_in_different_runs) for j in range(self.num_of_clusters)}
        
        ID_index = 0
        cluster_index = list(first_results[0]).index('clusters')
        n_data_index = list(first_results[0]).index('Num of data')
            
        
        for run_index in range(n_iterations):
            
            if run_index == 0:
                clustered_data_sorted = first_results
            else:
                if self.clustering_method  == 'KMeans':
                    clusterer = KMeans(n_clusters = self.num_of_clusters, init='k-means++', n_init=20, max_iter = 1000)
                    clusters = clusterer.fit(self.Kmean_input_data)
                    
                elif self.clustering_method  == 'KModes':
                    clusterer = KModes(n_clusters=self.num_of_clusters, init='Huang', n_init=50, verbose=False)
                    clusters = clusterer.fit_predict(self.Kmodes_input_data)
                    
                elif self.clustering_method  == 'KPrototypes':
                    clusterer = KPrototypes(n_clusters=self.num_of_clusters, init='Cao', n_init=50, verbose=False)
                    clusters = clusterer.fit_predict(self.Kprototypes_input_data, categorical=self.cateoricals_indices)
                    
                cluster_column = list(map(int, clusterer.labels_+1))
                cluster_column.insert(0, 'clusters')
                clustered_data = np.column_stack((self.data_for_clustering, cluster_column))
                cluster_index = list(clustered_data[0]).index('clusters')
                clustered_data_sorted = sorted(clustered_data[1:], key=lambda x: x[cluster_index])
                clustered_data_sorted = np.vstack((clustered_data[0], clustered_data_sorted))
                
                self.list_of_IDs = clustered_data_sorted[1:, 0]
                num_of_data_column = self.adding_num_of_data()
                
                clustered_data_sorted = np.column_stack((clustered_data_sorted, num_of_data_column))
                
                    
            
            ID_column = clustered_data_sorted[1:,ID_index]
            cluster_column = clustered_data_sorted[1:, cluster_index]
            n_data_column = clustered_data_sorted[1:, n_data_index]
            
            ID_cluster_array = np.column_stack((ID_column, cluster_column, n_data_column))
            
            divided_clusters ={i:[] for i in range(self.num_of_clusters)}
            for row in ID_cluster_array:
                cluster_as_key = int(row[1])-1
                divided_clusters[cluster_as_key] = np.append(divided_clusters[cluster_as_key], row[0])
        
            if run_index == 0:    
                first_clustering_results = dict(divided_clusters)
                for w in range(self.num_of_clusters):
                    clusters_in_different_runs[w][0] = first_clustering_results[w]
            else:
                next_clustering_results = dict(divided_clusters)
                
                for key_f, value_f in first_clustering_results.items():
                    max_n_common_IDs = 0
                    
                    for key_n, value_n in next_clustering_results.items():
                        common_IDs = list(set(value_f).intersection(value_n))
                        if max_n_common_IDs < len(common_IDs):
                            max_n_common_IDs = len(common_IDs)
                            value_of_max_common_IDs = value_n
                        
                    clusters_in_different_runs[key_f][run_index] = value_of_max_common_IDs
    
    ## Saving grouped clusters of different runs in an Excel file:       
            for i in range(self.num_of_clusters):
                test = clusters_in_different_runs[i][run_index]  
                
                max_column = len(test)
                for column in range(1, max_column+1):
                    _ = ws[i].cell(column=column, row=run_index+1, value = test[column-1])
            wb_clusters.save(filename = dest_filename0)
    
    
    ## checking the most occuring members of each cluster
        representative_IDs =[]
        for cluster_index in range(self.num_of_clusters):
            repeated_IDs = []
            for key,value in clusters_in_different_runs[cluster_index].items():
                repeated_IDs = np.append(repeated_IDs, value)
            temp_dic = Counter(repeated_IDs)
            temp_array = np.asarray(list((map(list, temp_dic.items()))))
            
            selected_IDs = [row[0] for row in temp_array if int(row[1]) >= n_iterations-1]
    
            selected_IDs_with_n_data = [row for row in ID_cluster_array if row[0] in selected_IDs]
            selected_ID = sorted(selected_IDs_with_n_data, key=lambda x: x[2])[len(selected_IDs_with_n_data)-1][0]
            representative_IDs.append(selected_ID)
        
    
    ## saving complete data of representative patients in an Excel file:
        representatives_data = self.data_set[0]
        data = [row for row in self.data_set if row[0] in representative_IDs]
        representatives_data = np.vstack((representatives_data, data))
        
        max_row = len(representatives_data)
        max_column = len(representatives_data[0]) 
        for row in range(1, max_row+1):
            for column in range(1, max_column+1):
                _ = ws_representatives.cell(column=column, row=row, value = representatives_data[row-1][column-1])
        wb_representatives.save(filename = dest_filename1)









