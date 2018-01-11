# -*- coding: utf-8 -*-

from openpyxl import Workbook
import scipy.stats as stats

import utilities as tools

import numpy as np







class correlation_analysis():
    
    def __init__(self, data_set=[], attributes=[], outcome_measures=[],  ordinals=[], categoricals=[]):
        
        self.data_set = data_set
        self.attributes = attributes
        self.ordinals = ordinals
        self.categoricals = categoricals
        self.outcome_measures = outcome_measures
        
     
    
    def filter_data(self, filtering_attribute='', filtering_value=0):
        self.data_set = tools.filter_for_attribute(data_set=self.data_set, filtering_attribute=filtering_attribute, filtering_value=filtering_value)
        
    
    def group_data(self):
        self.data_set = tools.group_data_by_attributes(data_set=self.data_set, grouping_attributes=self.attributes)
        
        
    def normalize_data(self):
        self.data_set = tools.normalize_data(self.data_set, self.categoricals)
        
    
    def assess_ordinal_correlation (self):
       
        self.correlation_table = np.hstack(('**ORDINAL CORRELATION**', self.ordinals))
        self.pValue_table = np.hstack(('**ORDINAL P VALUE**', self.ordinals))
        
        n_attributes = len(self.attributes)
        
        for outcome in self.outcome_measures:
            
            correlation_column = outcome
            pValue_column = outcome        
            outcome_index = list(self.data_set[0]).index(outcome) 
     
            for attribute in self.ordinals:
                if attribute != outcome:
                    attribute_index = list(self.data_set[0]).index(attribute)
               
                    if attribute_index < n_attributes:
                        paired_attributes_data = np.column_stack([self.data_set[:,attribute_index], self.data_set[:,outcome_index]])
                        completed_pairs = tools.del_incomplete_data (paired_attributes_data)
                        attribute_data = completed_pairs[1:,0]
                        outcome_data = completed_pairs[1:,1]
                                      
                    tau, p_value = stats.kendalltau(outcome_data, attribute_data)
                    
                    tau = round(abs(tau), 3)
                    p_value = round(p_value, 4)
                    correlation_column = np.hstack((correlation_column, tau))
                    pValue_column = np.hstack((pValue_column, p_value))
            
            self.correlation_table = np.column_stack((self.correlation_table, correlation_column))
            self.pValue_table = np.column_stack((self.pValue_table, pValue_column))
      
        return (self.correlation_table, self.pValue_table)
                    
    
    
    
    
    
    
    
    def assess_categorical_correlation(self):
        
        categorical_correlation_table = np.hstack(('**categorical CORRELATION**', self.categoricals))
        categorical_pValue_table = np.hstack(('**categorical P VALUE**', self.categoricals))
        
        for outcome in self.outcome_measures:
            
            categorical_correlation_column = outcome
            categorical_pValue_column = outcome
            outcome_index = list(self.data_set[0]).index(outcome)
            
            
            for attribute in self.categoricals:
                
                attribute_index = list(self.data_set[0]).index(attribute)
    
                paired_attributes = np.column_stack((self.data_set[:, attribute_index],
                                                                 self.data_set[:, outcome_index]))
                completed_pairs = tools.del_incomplete_data (paired_attributes)
    
                outcome_unique_values = np.unique(completed_pairs[1:, 1])
                attribute_unique_values = np.unique(completed_pairs[1:, 0])
                
                observed_table = []
                for outcome_value in outcome_unique_values:
                    observed_column = []
                    for attribute_value in  attribute_unique_values:
                        counter = 0
                        for row in completed_pairs:
                            if row[0] == attribute_value and row[1] == outcome_value:
                                counter += 1
                        observed_column = np.hstack((observed_column, counter))
                        
                    if observed_table == []:
                        observed_table = observed_column
                    else:
                        observed_table = np.column_stack((observed_table, observed_column))
                 
                    
                test_statistic, p_value, DoF, expected = stats.chi2_contingency(observed= observed_table)
                
                test_statistic = round((test_statistic/100), 4)
                p_value = round(p_value, 4)
                categorical_correlation_column = np.hstack((categorical_correlation_column, test_statistic))
                categorical_pValue_column = np.hstack((categorical_pValue_column, p_value))
            
            categorical_correlation_table = np.column_stack((categorical_correlation_table, categorical_correlation_column))
            categorical_pValue_table = np.column_stack((categorical_pValue_table, categorical_pValue_column))
        
        self.correlation_table = np.vstack((self.correlation_table, categorical_correlation_table))
        self.pValue_table = np.vstack((self.pValue_table, categorical_pValue_table))    
    
    
    
    
    def put_results_in_excel(self):
        wb_correlation = Workbook()
        ws_correlation = wb_correlation.create_sheet("Correlation",0)
        ws_pValues = wb_correlation.create_sheet("pValues",1)
        dest_filename_correlation = './results/correlation_results.xlsx'
        
        max_row = len(self.correlation_table)
        max_column = len(self.correlation_table[0]) 
        for row in range(1, max_row+1):
            for column in range(1, max_column+1):
                _ = ws_correlation.cell(column=column, row=row, value = self.correlation_table[row-1][column-1])
        
        
        max_row = len(self.pValue_table)
        max_column = len(self.pValue_table[0]) 
        for row in range(1, max_row+1):
            for column in range(1, max_column+1):
                _ = ws_pValues.cell(column=column, row=row, value = self.pValue_table[row-1][column-1])
        
        wb_correlation.save(filename = dest_filename_correlation)
                
        print("\n\nCheck the folder 'results' for correlation results in an Excel file nammed \"correlation_results\"!\n\n")
            
           


    def run(self):
        if self.ordinals:
            self.assess_ordinal_correlation()
        if self.categoricals:
            self.assess_categorical_correlation()
        
        self.put_results_in_excel()









