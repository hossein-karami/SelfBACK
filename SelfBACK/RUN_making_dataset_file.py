# -*- coding: utf-8 -*-
import copy as cp
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook




##  reading and executing configuration file for different data sets

dataset_input = input('enter the dataset which you want ot analyze it (f1: fysioprim 01/2017,  f8: fysioprim 08/2017,  d:Dphacto): ')
if dataset_input == 'f1':
    file = open ('./config_files/fysioprim_012017_configfile.py')
elif dataset_input == 'f8':
    file = open ('./config_files/fysioprim_082017_configfile.py')
elif dataset_input == 'd':
    file = open ('./config_files/dphacto_configfile.py')


file_lines = file.readlines()
for line in file_lines:
    exec(line)


numpy_dataset_destination = numpy_dataset_destination_address_and_filename    
xlFile_addr = excel_dataset_address 
wb = load_workbook(xlFile_addr)
sheet = wb.get_sheet_by_name(excel_sheet_name)

attribute_filename = './attributes.py'
ID_attribute = ID_attribute
add_BMI = if_add_BMI
height_attr = height_attribute
weight_attr = weight_attribute
manipulated_outcomes = manipulated_outcomes


def make_filtered_dataset(sheet, ordinal_attribute_set, categorical_attribute_set, outcome_measures):
    
    data = list(sheet.values)
    data_array = np.asarray(data)
    
    temp_array = cp.deepcopy(data_array)

## adding BMI to the dataset if needed    
    if add_BMI:
        height_index = list(temp_array[0]).index(height_attr)
        weight_index = list(temp_array[0]).index(weight_attr)
        BMI_column = []
        for row in temp_array[1:]:
            if row[height_index] != None and row[weight_index] != None:
                
                BMI = row[weight_index] / ((row[height_index]/100)**2)
                BMI_column.append(BMI)
            else:
                BMI_column.append(None)
                
        BMI_column.insert(0, 'BMI') 
        temp_array = np.column_stack((temp_array, BMI_column))
        ordinal_attribute_set.append('BMI')
        
    attribute_set_for_analysis = ordinal_attribute_set + categorical_attribute_set + outcome_measures

##  Removing columns and attributes not used in the analysis   
    for attribute in data_array[0]:
        if attribute == None or attribute not in attribute_set_for_analysis and attribute != ID_attribute:
            none_index = list(temp_array[0]).index(attribute)
            temp_array = np.delete(temp_array, none_index, axis=1)
    
    
    
    

##  Removing repeated instances
    ID = 0
    ID_index = list(temp_array[0]).index(ID_attribute)
    row_index =0
    for row in temp_array:
        if row[ID_index] !=ID:
            ID = row[ID_index]
        else:
            temp_array = np.delete(temp_array, row_index, axis=0)
            row_index -=1
        
        row_index += 1
    
    ID_index = list(temp_array[0]).index(ID_attribute)
    if ID_index != 0:
        ID_column = temp_array[:, ID_index]
        temp_array_2 = np.delete(temp_array, ID_index, axis=1)
        temp_array = np.column_stack((ID_column, temp_array_2))





##  manipulate the attributes to make a set of outcome measures
    if manipulated_outcomes:
        n_outcomes = len(outcome_measures) - 3
        outcome_titles = ['outcome'+str(i+1) for i in range(n_outcomes)]
        
        base_attr1 = 'OA_sms_resp_Week_00_Q2'
        base_attr2 = 'OA_sms_resp_Week_12_Q5'
        base_attr3 = 'OA_sms_resp_Week_12_Q6'
        
        Q2_attributes = [attr for attr in outcome_measures if 'Q2' in attr ]
        Q5_attributes = [attr for attr in outcome_measures if 'Q5' in attr ]
        Q6_attributes = [attr for attr in outcome_measures if 'Q6' in attr ]
        
        i = 0        
        for attribute in temp_array[0]:
            Q2_based_index = list(temp_array[0]).index(base_attr1) 
                        
            if attribute in Q2_attributes and attribute != base_attr1:
                Q2_outcome_column = [outcome_titles[i]]
                i += 1
                attr_index = list(temp_array[0]).index(attribute)
                
                for row in temp_array[1:]:
                    outcome_value = np.sign(row[Q2_based_index]-row[attr_index]) if None not in (row[Q2_based_index], row[attr_index]) else None
                    Q2_outcome_column.append(outcome_value)
                
                temp_array = np.column_stack((temp_array, Q2_outcome_column))
            
        for attribute in temp_array[0]:
            Q5_based_index = list(temp_array[0]).index(base_attr2) 
            
            if attribute in Q5_attributes and attribute != base_attr2:
                Q5_outcome_column = [outcome_titles[i]]
                i += 1
                attr_index = list(temp_array[0]).index(attribute)
                for row in temp_array[1:]:
                    outcome_value = np.sign(row[Q5_based_index]-row[attr_index]) if None not in (row[Q5_based_index], row[attr_index]) else None
                    Q5_outcome_column.append(outcome_value)
                
                temp_array = np.column_stack((temp_array, Q5_outcome_column))
        
        for attribute in temp_array[0]: 
            Q6_based_index = list(temp_array[0]).index(base_attr3) 
            
            if attribute in Q6_attributes and attribute != base_attr3:
                Q6_outcome_column = [outcome_titles[i]]
                i += 1
                attr_index = list(temp_array[0]).index(attribute)
                for row in temp_array[1:]:
                    outcome_value = np.sign(row[Q6_based_index]-row[attr_index]) if None not in (row[Q6_based_index], row[attr_index]) else None
                    Q6_outcome_column.append(outcome_value)
                
                temp_array = np.column_stack((temp_array, Q6_outcome_column))
        
        
        for attribute in outcome_measures:
            delete_index = list(temp_array[0]).index(attribute)
            temp_array = np.delete(temp_array, delete_index, axis=1)
            
        outcome_measures = outcome_titles
    


##  Writing the attribute python file
    attributes_file = open(attribute_filename, 'w')
    attributes_file.write('# -*- coding: utf-8 -*- \n\n\n')
    attributes_file.write('ordinal_attribute_set = [')
    for attribute in ordinal_attribute_set:
        attributes_file.write("'" + str(attribute) + "'" + ', ')
    attributes_file.write(']\n\n\n')
    
    attributes_file.write('categorical_attribute_set = [')
    for attribute in categorical_attribute_set:
        attributes_file.write("'" + str(attribute) + "'" + ', ')
    attributes_file.write(']\n\n\n')
    
    attributes_file.write('outcomes_measures = [')
    for attribute in outcome_measures:
        attributes_file.write("'" + str(attribute) + "'" + ', ')
    attributes_file.write(']\n\n\n')
      
    attributes_file.close()
    np.save(numpy_dataset_destination, temp_array)
    


##  making an excel file from the dataset which is made for analysis, just to check (if needed)    
#    wb_data = Workbook()
#    ws = wb_data.create_sheet("dataset",0)
#    
#    dest_filename = '../datasets/dataset_used_for_analysis.xlsx'
#    
#    max_row = len(temp_array)
#    max_column = len(temp_array[0]) 
#    for row in range(1, max_row+1):
#        for column in range(1, max_column+1):
#            _ = ws.cell(column=column, row=row, value = temp_array[row-1][column-1])
#    
#    wb_data.save(filename = dest_filename)
#    print('You can check the dataset made, in the dataset folder')




 
    
    
def make_attribute_file(sheet, attribute_filename)  :
    
    ordinal_attribute_set =[]
    categorical_attribute_set = []
    outcome_generators = []
    ordinal_attributes_color = ''
    categorical_attributes_color = ''
    outcome_attributes_color = ''
    
    
    for item in sheet[1]:
        if item.value == first_colored_ordinal_attribute:
            ordinal_attributes_color = item.fill.start_color.index
        elif item.value == first_colored_categorical_attribute:
            categorical_attributes_color = item.fill.start_color.index
        elif item.value == first_colored_outcome_attribute:
            outcome_attributes_color = item.fill.start_color.index
              
        
        if item.fill.start_color.index == ordinal_attributes_color:
            ordinal_attribute_set.append(item.value) 
        elif item.fill.start_color.index == categorical_attributes_color:
            categorical_attribute_set.append(item.value)
        elif item.fill.start_color.index == outcome_attributes_color:
            outcome_generators.append(item.value)
    
    return ordinal_attribute_set, categorical_attribute_set, outcome_generators






ordinal_attribute_set, categorical_attribute_set, outcome_generators = make_attribute_file(sheet, attribute_filename)
make_filtered_dataset(sheet, ordinal_attribute_set, categorical_attribute_set, outcome_generators)

print('\nDataset numpy file is ready to use and is saved in dastination address mentioned in config file.')
