import numpy as np
from openpyxl import Workbook
from imputation_methods import *
from attributes import *
from cluster_analysis_class import *
import utilities as tools


dataset_input = input('enter the dataset which you want ot analyze it (f1: fysioprim 01/2017,  f8: fysioprim 08/2017,  d:Dphacto): ')
if dataset_input == 'f1':
    file = open ('./config_files/fysioprim_012017_configfile.py')
elif dataset_input == 'f8':
    file = open ('./config_files/fysioprim_082017_configfile.py')
elif dataset_input == 'd':
    file = open ('./config_files/dphacto_configfile.py')


file_lines = file.readlines()
for line in file_lines:
    exec(line)
        

dataset_array = np.load(numpy_dataset_destination_address_and_filename+'.npy')



attributes = ordinal_attribute_set + categorical_attribute_set + outcomes_measures
                              
filtered_data =  tools.filter_for_attribute(data_set=dataset_array, attributes=attributes, filtering_attribute=filtering_attribute, filtering_value=filtering_value) if if_filter else dataset_array

normalized_data = tools.normalize_data(data_set=filtered_data, categoricals=categorical_attribute_set) if if_normalize else filtered_data
reduced_data_for_clustering = tools.group_data_by_attributes(data_set=normalized_data, grouping_attributes=attributes_for_clustering)





#____________ Different methods of dealing with incomplete data, choose one of them
if imputation_method == 'REMOVE':
    completed_data = del_incomplete_data (reduced_data_for_clustering)
elif imputation_method == 'MEANS':
    completed_data = completing_by_mean_value (reduced_data_for_clustering)
elif imputation_method == 'KNN':
    completed_data = completing_by_KNN(normalized_data, reduced_data_for_clustering, categorical_attribute_set)    





#____________ Cluster analysis and box plots
cluster_analizer = dataset_clustering(data_set=dataset_array, data_for_clustering=completed_data, attributes_for_clustering=attributes_for_clustering, categoricals=categorical_attribute_set,
                                      ordinals=ordinal_attribute_set, n_test_clusters=n_test_clusters, clustering_method=clustering_method)
if clustering_method == 'KModes':
    clustered_data_sorted = cluster_analizer.run_KModes()
elif clustering_method == 'KMeans':
    clustered_data_sorted = cluster_analizer.run_KMeans()
elif clustering_method == 'KPrototypes':
    clustered_data_sorted = cluster_analizer.run_KPrototypes()
    

if if_representatives: cluster_analizer.select_representative_IDs(first_results=clustered_data_sorted, n_iterations=6)
if if_boxplots: cluster_analizer.draw_boxplots()




#____________ Saving different datasets in an Excel file, named "Processed_Datasets":

wb_Data = Workbook()
ws_filtered = wb_Data.create_sheet("Filtered",0)
ws_normalized = wb_Data.create_sheet("Normalized",1)
ws_reduced = wb_Data.create_sheet("Reduced",2)
ws_completed = wb_Data.create_sheet("Completed",3)
ws_clustered = wb_Data.create_sheet("Clustered",4)
dest_filename = './results/Processed_Datasets.xlsx'


worksheets_name =[ws_filtered, ws_normalized, ws_reduced, ws_completed, ws_clustered]
temp_dic = {0:filtered_data, 1:normalized_data, 2:reduced_data_for_clustering, 3:completed_data, 4:clustered_data_sorted}

for i in range(len(worksheets_name)):
    
    max_row = len(temp_dic[i])
    max_column = len(temp_dic[i][0]) 
    for row in range(1, max_row+1):
        for column in range(1, max_column+1):
            _ = worksheets_name[i].cell(column=column, row=row, value = temp_dic[i][row-1][column-1])
    wb_Data.save(filename = dest_filename)  


 